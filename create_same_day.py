import pandas as pd
import openpyxl
import numpy as np
def write_2excel():
    wb=openpyxl.load_workbook('report.xlsx')
    sheet1=wb['Sheet1']
    def header():
        #start row=1, col =1-15
        sheet1.cell(row=1,column=1).value='shopid'
        sheet1.cell(row=1, column=2).value = 'username'
        sheet1.cell(row=1, column=3).value = 'seller_type'
        sheet1.cell(row=1, column=4).value = 'main_category'
        sheet1.cell(row=1, column=5).value = 'seller_phone'
        sheet1.cell(row=1, column=6).value = 'seller_address(province)'
        #>320
        sheet1.cell(row=1, column=7).value = 'order'
        sheet1.cell(row=1, column=8).value = 'GMV'
        #>400
        sheet1.cell(row=1, column=9).value = 'order'
        sheet1.cell(row=1, column=10).value = 'GMV'
        #>1000
        sheet1.cell(row=1, column=11).value = 'order'
        sheet1.cell(row=1, column=12).value = 'GMV'
        #>1600
        sheet1.cell(row=1, column=13).value = 'order'
        sheet1.cell(row=1, column=14).value = 'GMV'
        #conclude overall
        sheet1.cell(row=1, column=15).value = 'total_order'
        sheet1.cell(row=1, column=16).value = 'GMV'
    #seller detail starts row=2, col =1,2,3,4,5
    # df_detail=pd.DataFrame(pd.read_csv('seller_detail.csv'))
    # df_detail.sort_values(by='shopid')
    # # df_detail=df_detail.set_index('shopid')
    # row=2
    # print(df_detail.head())
    header()
    df_shopid = pd.DataFrame(pd.read_excel('shopid.xlsx'))
    df_shopid = df_shopid.set_index('shopid')
    df_shopid['a'] = 0
    #<320 start row=2,  col =7,8
    df_320=pd.read_csv('order_320.csv')
    df_320=df_320.set_index('shopid')
    df_join=df_shopid.join(df_320).fillna(0)
    row=2
    row_pd=0
    for n in range(len(df_join)):
        sheet1.cell(row=row, column=7).value =df_join.iloc[row_pd, 1]
        sheet1.cell(row=row, column=8).value = df_join.iloc[row_pd, 2]
        row+=1
        row_pd+=1
    #>400 start row=2, col=9,10
    df_400=pd.read_csv('order_400.csv')
    df_400=df_400.set_index('shopid')
    df_join = df_shopid.join(df_400).fillna(0)
    row = 2
    row_pd = 0
    for n in range(len(df_join)):
        sheet1.cell(row=row, column=9).value = df_join.iloc[row_pd, 1]
        sheet1.cell(row=row, column=10).value = df_join.iloc[row_pd, 2]
        row += 1
        row_pd += 1
    #>1000 start row=2, column=11,12
    df_1000 = pd.read_csv('order_1000.csv')
    df_1000 = df_1000.set_index('shopid')
    df_join = df_shopid.join(df_1000).fillna(0)
    row = 2
    row_pd = 0
    for n in range(len(df_join)):
        sheet1.cell(row=row, column=11).value = df_join.iloc[row_pd, 1]
        sheet1.cell(row=row, column=12).value = df_join.iloc[row_pd, 2]
        row += 1
        row_pd += 1
    # >1600 start row=2, column=13,14
    df_1600 = pd.read_csv('order_1600.csv')
    df_1600 = df_1600.set_index('shopid')
    df_join = df_shopid.join(df_1600).fillna(0)
    row = 2
    row_pd = 0
    for n in range(len(df_join)):
        sheet1.cell(row=row, column=13).value = df_join.iloc[row_pd, 1]
        sheet1.cell(row=row, column=14).value = df_join.iloc[row_pd, 2]
        row += 1
        row_pd += 1
    #overall start row=2, col=15,16
    df_overall = pd.read_csv('order_overall.csv')
    df_overall = df_overall.set_index('shopid')
    df_join = df_shopid.join(df_overall).fillna(0)
    row = 2
    row_pd = 0
    for n in range(len(df_join)):
        sheet1.cell(row=row, column=15).value = df_join.iloc[row_pd, 1]
        sheet1.cell(row=row, column=16).value = df_join.iloc[row_pd, 2]
        row += 1
        row_pd += 1
    wb.save('report.xlsx')
write_2excel()

