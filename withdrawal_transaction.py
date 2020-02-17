import pandas as pd
import openpyxl

def withdrawal_transaction():
    wb=openpyxl.load_workbook('withdrawal_order_transaction_report.xlsx')
    sheet1=wb['Withdrawal']
    df_bank=[]
    dict_bank_col={}
    def set_header():
        df_bankName=pd.read_csv('bank_for_withdrawal.csv').sort_values(by='bank')
        nonlocal df_bank
        df_bank=df_bankName
        #Time label
        sheet1.column_dimensions['A'].width=10
        sheet1.cell(1,1).value='Time'
        #Bank label
        sheet1.column_dimensions['B'].width = 10
        sheet1.cell(1,2).value='Bank'
        row=1
        row_pd=0
        for col in range(3,len(df_bankName)+2):
            sheet1.cell(row, col).value=df_bankName.iloc[row_pd,0]
            nonlocal dict_bank_col
            dict_bank_col[df_bankName.iloc[row_pd,0]]=col
            row_pd+=1
        wb.save('withdrawal_order_transaction_report.xlsx')
    def input_values():
        df_trans_withdrawal=pd.read_csv('transaction_withdrawal.csv')
        unique_time=pd.Series(df_trans_withdrawal['ctime'].unique())
        start=0
        i = 0
        row_for_trans =2
        row_for_amount=3
        row = 0
        while True:
            row_pd=0
            while True:
                try:
                    if df_trans_withdrawal.iloc[row,0] == unique_time[i]:
                        row+=1
                        continue
                    else:
                        sheet1.cell(row_for_trans,1).value = unique_time[i]
                        sheet1.cell(row_for_trans,2).value = 'total transaction'
                        sheet1.cell(row_for_amount,2).value= 'total amount'
                        i+=1
                        row_pd=row
                        break
                except Exception: break
            # print(row_pd)
            same_date= df_trans_withdrawal.iloc[start:row_pd, :]
            if same_date.empty: break
            start=row_pd
            # print(same_date)
            index_for_same_date=0
            l=0
            # adapt sheet width
            for item in range(ord('C'), ord('R')+1):
                sheet1.column_dimensions[chr(item)].width=25
            #put total transaction, amount that corresponding to its bank
            while True:
                try:
                    if [same_date.iloc[index_for_same_date,1]] in df_bank.values.tolist():
                        #total transaction
                        sheet1.cell(row_for_trans, dict_bank_col[same_date.iloc[index_for_same_date,1]]).value=float(same_date.iloc[index_for_same_date,2])
                        #total amount
                        sheet1.cell(row_for_amount, dict_bank_col[same_date.iloc[index_for_same_date,1]]).value=float(same_date.iloc[index_for_same_date,3])
                        l += 1
                        index_for_same_date+=1
                    else:
                        # # total transaction
                        # sheet1.cell(row_for_trans, dict_bank_col[same_date.iloc[index_for_same_date, 1]]).value = ''
                        # # total amount
                        # sheet1.cell(row_for_amount, dict_bank_col[same_date.iloc[index_for_same_date, 1]]).value = ''
                        index_for_same_date=l
                        l+=1
                        continue
                except Exception : break
            row_for_amount+=2
            row_for_trans+=2
        wb.save('withdrawal_order_transaction_report.xlsx')
    set_header()
    input_values()
withdrawal_transaction()