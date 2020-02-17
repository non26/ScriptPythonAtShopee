# import pandas as pd
# import openpyxl
# def writeMS_2excel():
#     wb=openpyxl.load_workbook('report.xlsx')
#     ms=wb['MS']
#     # ado row=2. col=5,6,7
#     col = 5
#     for name in ('_june', '_july', '_aug'):
#         x = pd.DataFrame(pd.read_csv('ms_detail.csv').iloc[:, 0])
#         df = pd.read_csv(f"MS_ado{name}.csv")
#         z = x.join(df, lsuffix='_l', rsuffix='_r')
#         z.sort_values
#         z = z.fillna(0)
#         # print(z)
#         row = 2
#         for i in range(len(z)):
#             ms.cell(row=row, column=col).value = z.iloc[i, 2]
#             row += 1
#         col += 1
#     # nfr row=2 col=8,9,10
#     col = 8
#     for name in ('_june', '_july', '_aug'):
#         x = pd.DataFrame(pd.read_csv('ms_detail.csv').iloc[:, 0])
#         df = pd.read_csv(f"MS_nfr{name}.csv")
#         z = x.join(df, lsuffix='_l', rsuffix='_r')
#         z.sort_values
#         z = z.fillna('0')
#         # print(z)
#         row = 2
#         for i in range(len(z)):
#             ms.cell(row=row, column=col).value = z.iloc[i, 2]
#             row += 1
#         col += 1
#     # lsr row=2 col 11,12,13
#     col = 11
#     for name in ('_june', '_july', '_aug'):
#         x = pd.DataFrame(pd.read_csv('ms_detail.csv').iloc[:, 0])
#         df = pd.read_csv(f"MS_lsr{name}.csv")
#         z = x.join(df, lsuffix='_l', rsuffix='_r')
#         z.sort_values
#         z = z.fillna(0)
#         # print(z)
#         row = 2
#         for i in range(len(z)):
#             ms.cell(row=row, column=col).value = z.iloc[i, 2]
#             row += 1
#         col += 1
#     wb.save('report.xlsx')
# def writeOS_2excel():
#     wb = openpyxl.load_workbook('report.xlsx')
#     os = wb['OS']
#     # ado row=2. col=5,6,7
#     col = 5
#     for name in ('_june', '_july', '_aug'):
#         x = pd.DataFrame(pd.read_csv('os_detail.csv').iloc[:, 0])
#         df = pd.read_csv(f"OS_ado{name}.csv")
#         z = x.join(df, lsuffix='_l', rsuffix='_r')
#         z.sort_values
#         z = z.fillna(0)
#         # print(z)
#         row = 2
#         for i in range(len(z)):
#             os.cell(row=row, column=col).value = z.iloc[i, 2]
#             row += 1
#         col += 1
#     # nfr row=2, col=8,9,10
#     col = 8
#     for name in ('_june', '_july', '_aug'):
#         x = pd.DataFrame(pd.read_csv('os_detail.csv').iloc[:, 0])
#         df = pd.read_csv(f"OS_nfr{name}.csv")
#         z = x.join(df, lsuffix='_l', rsuffix='_r')
#         z.sort_values
#         z = z.fillna(0)
#         # print(z)
#         row = 2
#         for i in range(len(z)):
#             os.cell(row=row, column=col).value = z.iloc[i, 2]
#             row += 1
#         col += 1
#     # lsr row=2, col =11,12,13
#     col = 11
#     for name in ('_june', '_july', '_aug'):
#         x = pd.DataFrame(pd.read_csv('os_detail.csv').iloc[:, 0])
#         df = pd.read_csv(f"OS_nfr{name}.csv")
#         z = x.join(df, lsuffix='_l', rsuffix='_r')
#         z.sort_values
#         z = z.fillna(0)
#         # print(z)
#         row = 2
#         for i in range(len(z)):
#             os.cell(row=row, column=col).value = z.iloc[i, 2]
#             row += 1
#         col += 1
#     wb.save('report.xlsx')
# writeMS_2excel()
# writeOS_2excel()
import pandas as pd
import openpyxl
from openpyxl.workbook import Workbook
def writeMS_2excel():
    wb=openpyxl.load_workbook('report_campaign.xlsx')
    sheet=wb['MS']
    #ado row=2, col 5,6,7,8
    col = 5
    for name in ('_6.6june','_27.6june', '_7.7july', '_8.8aug'):
        x = pd.DataFrame(pd.read_csv('ms_detail.csv').iloc[:, 0])
        df = pd.read_csv(f"MS_ado{name}.csv")
        x = x.set_index('shopid')
        df = df.set_index('shopid')
        z = x.join(df, lsuffix='_l', rsuffix='_r')
        # z.sort_values()
        z = z.fillna(0)
        # print(z)
        row = 2
        for i in range(len(z)):
            sheet.cell(row=row, column=col).value = z.iloc[i, 0]
            row += 1
        col += 1
    #nfr row=2, col=9,10,11
    col = 9
    for name in ('_june', '_july', '_aug'):
        x = pd.DataFrame(pd.read_csv('ms_detail.csv').iloc[:, 0])
        df = pd.read_csv(f"MS_nfr{name}.csv")
        x = x.set_index('shopid')
        df = df.set_index('shopid')
        z = x.join(df, lsuffix='_l', rsuffix='_r')
        # z.sort_values()
        z = z.fillna(0)
        # print(z)
        row = 2
        for i in range(len(z)):
            sheet.cell(row=row, column=col).value = z.iloc[i, 0]
            row += 1
        col += 1
    #lsr row= 2, col=12,13,14
    col = 12
    for name in ('_june', '_july', '_aug'):
        x = pd.DataFrame(pd.read_csv('ms_detail.csv').iloc[:, 0])
        df = pd.read_csv(f"MS_nfr{name}.csv")
        x = x.set_index('shopid')
        df = df.set_index('shopid')
        z = x.join(df, lsuffix='_l', rsuffix='_r')
        # z.sort_values()
        z = z.fillna(0)
        # print(z)
        row = 2
        for i in range(len(z)):
            sheet.cell(row=row, column=col).value = z.iloc[i, 0]
            row += 1
        col += 1
    wb.save('report_campaign.xlsx')
def writeOS_2excel():
    wb = openpyxl.load_workbook('report_campaign.xlsx')
    sheet = wb['OS']
    # ado row=2, col 5,6,7,8
    col = 5
    for name in ('_6.6june', '_27.6june', '_7.7july', '_8.8aug'):
        x = pd.DataFrame(pd.read_csv('os_detail.csv').iloc[:, 0])
        df = pd.read_csv(f"OS_ado{name}.csv")
        x = x.set_index('shopid')
        df = df.set_index('shopid')
        z = x.join(df, lsuffix='_l', rsuffix='_r')
        # z.sort_values()
        z = z.fillna(0)
        # print(z)
        row = 2
        for i in range(len(z)):
            sheet.cell(row=row, column=col).value = z.iloc[i, 0]
            row += 1
        col += 1
    # nfr row=2, col=9,10,11
    col = 9
    for name in ('_june', '_july', '_aug'):
        x = pd.DataFrame(pd.read_csv('os_detail.csv').iloc[:, 0])
        df = pd.read_csv(f"OS_nfr{name}.csv")
        x = x.set_index('shopid')
        df = df.set_index('shopid')
        z = x.join(df, lsuffix='_l', rsuffix='_r')
        # z.sort_values()
        z = z.fillna(0)
        # print(z)
        row = 2
        for i in range(len(z)):
            sheet.cell(row=row, column=col).value = z.iloc[i, 0]
            row += 1
        col += 1
    # lsr row= 2, col=12,13,14
    col = 12
    for name in ('_june', '_july', '_aug'):
        x = pd.DataFrame(pd.read_csv('os_detail.csv').iloc[:, 0])
        df = pd.read_csv(f"OS_nfr{name}.csv")
        x = x.set_index('shopid')
        df = df.set_index('shopid')
        z = x.join(df, lsuffix='_l', rsuffix='_r')
        # z.sort_values()
        z = z.fillna(0)
        # print(z)
        row = 2
        for i in range(len(z)):
            sheet.cell(row=row, column=col).value = z.iloc[i, 0]
            row += 1
        col += 1
    wb.save('report_campaign.xlsx')
writeMS_2excel()
writeOS_2excel()