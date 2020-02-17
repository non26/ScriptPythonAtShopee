import pandas as pd
import openpyxl
def write_month_2excel(seller):
    wb=openpyxl.load_workbook('report_v2.xlsx')
    sheet1=wb[f"{seller}"]
    print(f"seller={seller}")
    def header():
        #start row=3
        df=pd.read_excel('detail.xlsx', sheet_name=f"{seller}")
        row=3
        row_pd=0
        # shopid
        sheet1.cell(row=row - 1, column=1).value = 'shopid'
        # userid
        sheet1.cell(row=row - 1, column=2).value = 'userid'
        # username
        sheet1.cell(row=row - 1, column=3).value = 'username'
        # main_category
        sheet1.cell(row=row - 1, column=4).value = 'main_category'
        while True:
            try:
                # shopid
                sheet1.cell(row=row, column=1).value=df.iloc[row_pd, 0]
                # userid
                sheet1.cell(row=row, column=2).value = df.iloc[row_pd, 1]
                # username
                sheet1.cell(row=row, column=3).value = df.iloc[row_pd, 2]
                # main_category
                sheet1.cell(row=row, column=4).value = df.iloc[row_pd, 3]
                row+=1
                row_pd+=1
            except Exception: break
    print('header')
    header()
    def ado():
        col = 5
        for name in ('_june', '_july', '_aug'):
            x = pd.DataFrame(pd.read_excel('detail.xlsx', sheet_name=f"{seller}").iloc[:, 0])
            df = pd.read_csv(f"{seller}_ado{name}.csv")
            x = x.set_index('shopid')
            df = df.set_index('shopid')
            z = x.join(df, lsuffix='_l', rsuffix='_r')
            # z.sort_values
            z = z.fillna(0)
            # print(z)
            row = 3
            for i in range(len(z)):
                sheet1.cell(row=row, column=col).value = z.iloc[i, 0]
                row += 1
            col += 1
    print('ado')
    ado()
    def nfr():
        col = 8
        for name in ('_june', '_july', '_aug'):
            x = pd.DataFrame(pd.read_excel('detail.xlsx', sheet_name=f"{seller}").iloc[:, 0])
            df = pd.read_csv(f"{seller}_nfr{name}.csv")
            x = x.set_index('shopid')
            df = df.set_index('shopid')
            z = x.join(df, lsuffix='_l', rsuffix='_r')
            # z.sort_values()
            z = z.fillna('0')
            row = 3
            for i in range(len(z)):
                sheet1.cell(row=row, column=col).value = z.iloc[i, 0]
                row += 1
            col += 1
    print('nfr')
    nfr()
    def lsr():
        col = 11
        for name in ('_june', '_july', '_aug'):
            x = pd.DataFrame(pd.read_excel('detail.xlsx', sheet_name=f"{seller}").iloc[:, 0])
            df = pd.read_csv(f"{seller}_lsr{name}.csv")
            x = x.set_index('shopid')
            df = df.set_index('shopid')
            z = x.join(df, lsuffix='_l', rsuffix='_r')
            # z.sort_values
            z = z.fillna(0)
            # print(z)
            row = 3
            for i in range(len(z)):
                sheet1.cell(row=row, column=col).value = z.iloc[i, 0]
                row += 1
            col += 1
    print('lsr')
    lsr()
    wb.save('report_v2.xlsx')
for name in ('OS', 'MS'):
    write_month_2excel(name)