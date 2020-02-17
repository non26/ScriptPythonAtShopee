import pandas as pd
import openpyxl
from openpyxl.workbook import Workbook
def write_Week_2excel(seller):
    wb = openpyxl.load_workbook('report_v2.xlsx')
    sheet = wb[f"{seller}"]
    print(f"{seller}")
    date = ['2019-06-03', '2019-06-10', '2019-06-17', '2019-06-24',
            '2019-07-01', '2019-07-08', '2019-07-15', '2019-07-22', '2019-07-29',
            '2019-08-05', '2019-08-12', '2019-08-19', '2019-08-26', '2019-09-02', '2019-09-09', '2019-09-16']
    df2 = pd.read_excel('detail.xlsx', sheet_name=f"{seller}").iloc[:, [0, 1]]
    df2 = df2.set_index('shopid')
    col=14
    def write_parameter(parameter):
        #start col=14, row=3
        nonlocal col
        print('col=', col)
        sheet.cell(row=1, column=col).value = f"{parameter}"
        # col=14 # how dose this function remember col value for each call ?
        for day in date:
            file=f"{seller}_{parameter}_weekly_{day}.csv"
            df1=pd.read_csv(file)
            df1=df1.set_index('shopid')
            z = df2.join(df1, lsuffix='_l', rsuffix='_r').fillna(0)
            # z = z.fillna(0)
            row = 3
            for n in range(len(z)):
                sheet.cell(row=2, column=col).value = day
                sheet.cell(row=row, column=col).value = z.iloc[n, 1]
                row += 1
            col += 1
            wb.save('report_v2.xlsx')
    for value in ('ado', 'nfr', 'lsr'):
        print(value)
        write_parameter(value)
    def pnp():
        print('matching penalty points per week')
        nonlocal  col
        print('col=', col)
        sheet.cell(row=1, column=col).value = 'Penalty points per week'
        for day in date:
            name1 = f"TH {day} penalty points summary.xlsx"
            df1 = pd.read_excel(name1, sheet_name='This week performance').iloc[:, [0, 5, 6, 7]]
            df1 = df1.set_index('shopid')
            z = df2.join(df1, how='left').fillna(0)
            row = 3
            for n in range(len(z)):
                sheet.cell(row=2, column=col).value = day
                sheet.cell(row=row, column=col).value = z.iloc[n, 1]
                row+=1
            col+=1
        print('current penalty points')
        row = 3
        name=f"{seller}_current_penalty.csv"
        df1=pd.read_csv(name)
        df1 = df1.set_index('shopid')
        z = df2.join(df1, how='left').fillna(0)
        sheet.cell(row=2, column=col).value = 'current_penalty_points'
        for n in range(len(z)):
            sheet.cell(row=row, column=col).value = int(z.iloc[n, 1])
            row+=1
        print('finish')
        wb.save('report_v2.xlsx')
    pnp()
# for name in ('OS', 'MS'):
# for name in ('OS',):
for name in ('MS',):
    write_Week_2excel(name)






