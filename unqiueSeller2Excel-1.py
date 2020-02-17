import pandas as pd
import xlsxwriter
import openpyxl
week=list(range(18,32))
seller_type=sorted(['LT','OS','ST','MT'])
def write_week_section():
    df_withdrawalWeek=pd.read_csv('num_withdrawal_week.csv')
    # writer=pd.ExcelWriter('unique_seller_report.xlsx', engine='xlsxwriter')
    # df_withdrawalWeek.to_excel(writer, sheet_name='1', index=False)
    wb=openpyxl.load_workbook('unique_seller_report.xlsx')
    sheet1=wb['Sheet1']
    sheet1['A1'].value='total withdrawal'
    sheet1['B2'].value='week'
    # put week value to column 2
    i=0
    for row in range(3, 3+31-18+1, 1):
        sheet1.cell(row, 2).value=week[i]
        i+=1
    # put seller type to row 2
    i=0
    for column in range(1,len(seller_type)+1):
        sheet1.cell(2, column+2).value=seller_type[i]
        i+=1
    # put values in df_withdrawalWeek at seller_type and total_withdrawal
    # to Sheet1 sheet by their seller type
    i=0
    row=3
    s1=df_withdrawalWeek.iloc[:, 1: ]
    start=0
    while True:
        row_pd=0
        a1=s1.iloc[start:4+i, :].sort_values(by='seller_type')
        for column in range(3, 7):
            sheet1.cell(row, column).value=a1.iloc[row_pd, 1]
            row_pd+=1
        if 4+i == len(df_withdrawalWeek):
            break
        start+=4
        i+=4
        row +=1
        # if 4+i == len(df_withdrawalWeek):
        #     break
    #put total seller in dialy at column 7
    row_pd=0
    row=3
    sheet1.cell(2,7).value='total seller'
    df_totalSellerWeek=pd.read_csv('num_seller_withdrawal_week.csv')
    for row in range(3, 3+31-18+1):
        sheet1.cell(row, 7).value=df_totalSellerWeek.iloc[row_pd, 1]
        row+=1
        row_pd+=1
    wb.save('unique_seller_report.xlsx')
def write_manual_section():
    wb=openpyxl.load_workbook('unique_seller_report.xlsx')
    ws=wb.active
    # put labels at column 2 row 18
    ws.cell(18,2).value='Manual withdrawal'
    ws.cell(19,2).value='Monthly (May-Jul)'
    ws.cell(20,2).value='May'
    ws.cell(21,2).value='June'
    ws.cell(22,2).value='July'
    # put seller type at column 3 row 18
    column=3
    for index in range(1,len(seller_type)+1):
        ws.cell(18,column).value=seller_type[index-1]
        column += 1
    # put total withdrawal by its seller type at row 19 column 3
    df_manWithdrawal=pd.read_csv('num_withdrawal_manual.csv').sort_values(by='seller_type')
    column=3
    row_pd=0
    for index in range(len(df_manWithdrawal)):
        ws.cell(19, column+index).value=df_manWithdrawal.iloc[row_pd,1]
        row_pd+=1
    # put total withdrawal by its seller type and month at row 20, col 3
    df_manWithdrawalPerMonth=pd.read_csv('num_withdrawal_manual_perMonth.csv').iloc[:, 1:]
    row = 20
    column =3
    start=0
    l=0
    while True:
        a1=df_manWithdrawalPerMonth.iloc[start:4+l, :]
        for i in range(4):
            ws.cell(row,column+i ).value=a1.iloc[i, 1]
        row+=1
        start+=4
        l+=4
        if row == 23:
            break
    #put total seller to row 18, col 7
    df_sellerWithdrawal=pd.read_csv('num_seller_withdrawal_manual.csv')
    ws.cell(18,7).value='total seller'
    # put the number of seller at row 19 col 7
    ws.cell(19,7).value=df_sellerWithdrawal.iloc[0,0]
    wb.save('unique_seller_report.xlsx')
def wrtie_auto_section():
    wb = openpyxl.load_workbook('unique_seller_report.xlsx')
    ws = wb.active
    # put labels at column 2 row 24
    ws.cell(24, 2).value = 'Auto withdrawal'
    ws.cell(25, 2).value = 'Monthly (May-Jul)'
    ws.cell(26, 2).value = 'May'
    ws.cell(27, 2).value = 'June'
    ws.cell(28, 2).value = 'July'
    # put seller type at column 3 row 24
    column = 3
    for index in range(1, len(seller_type) + 1):
        ws.cell(24, column).value = seller_type[index - 1]
        column += 1
    # put total withdrawal by its seller type at row 25 column 3
    df_manWithdrawal = pd.read_csv('num_withdrawal_auto.csv').sort_values(by='seller_type')
    column = 3
    row_pd = 0
    for index in range(len(df_manWithdrawal)):
        ws.cell(25, column + index).value = df_manWithdrawal.iloc[row_pd, 1]
        row_pd += 1
    # put total withdrawal by its seller type and month at row 26, col 3
    df_manWithdrawalPerMonth = pd.read_csv('num_withdrawal_auto_perMonth.csv').iloc[:, 1:]
    row = 26
    column = 3
    start = 0
    l = 0
    while True:
        a1 = df_manWithdrawalPerMonth.iloc[start:4 + l, :]
        for i in range(4):
            ws.cell(row, column + i).value = a1.iloc[i, 1]
        row += 1
        start += 4
        l += 4
        if row == 24+5:
            break
    # put total seller to row 24, col 7
    df_sellerWithdrawal = pd.read_csv('num_seller_withdrawal_auto.csv')
    ws.cell(24, 7).value = 'total seller'
    # put the number of seller at row 25 col 7
    ws.cell(25, 7).value = df_sellerWithdrawal.iloc[0, 0]
    wb.save('unique_seller_report.xlsx')
def write_manAuto_section():
    wb = openpyxl.load_workbook('unique_seller_report.xlsx')
    ws = wb.active
    # put labels at column 2 row 30
    ws.cell(30, 2).value = 'Auto withdrawal'
    ws.cell(31, 2).value = 'Monthly (May-Jul)'
    ws.cell(32, 2).value = 'May'
    ws.cell(33, 2).value = 'June'
    ws.cell(34, 2).value = 'July'
    # put seller type at column 3 row 30
    column = 3
    for index in range(1, len(seller_type) + 1):
        ws.cell(30, column).value = seller_type[index - 1]
        column += 1
    # put total withdrawal by its seller type at row 31 column 3
    df_manWithdrawal = pd.read_csv('num_withdrawal_manauto.csv').sort_values(by='seller_type')
    column = 3
    row_pd = 0
    for index in range(len(df_manWithdrawal)):
        ws.cell(31, column + index).value = df_manWithdrawal.iloc[row_pd, 1]
        row_pd += 1
    # put total withdrawal by its seller type and month at row 32, col 3
    df_manWithdrawalPerMonth = pd.read_csv('num_withdrawal_manauto_perMonth.csv').iloc[:, 1:]
    row = 32
    column = 3
    start = 0
    l = 0
    while True:
        a1 = df_manWithdrawalPerMonth.iloc[start:4 + l, :]
        for i in range(4):
            ws.cell(row, column + i).value = a1.iloc[i, 1]
        row += 1
        start += 4
        l += 4
        if row == 30 + 5:
            break
    # put total seller to row 30, col 7
    df_sellerWithdrawal = pd.read_csv('num_seller_withdrawal_manauto.csv')
    ws.cell(30, 7).value = 'total seller'
    # put the number of seller at row 31 col 7
    ws.cell(31, 7).value = df_sellerWithdrawal.iloc[0, 0]
    wb.save('unique_seller_report.xlsx')
    
write_week_section()
write_manual_section()
wrtie_auto_section()
write_manAuto_section()




