import pandas as pd
import openpyxl
import datetime
import numpy as np
def write_to_excel():
    wb=openpyxl.load_workbook('report.xlsx')
    sheet1=wb['Sheet1']
    def switch_column_re3():
        # re3 = pd.read_csv('order_backlog_part3.csv') # it's cancel.csv
        re3 = pd.read_csv('cancel.csv')
        gb=re3.groupby('cancel_user')
        gb_re3=pd.DataFrame()
        for name, group in gb:
            if gb_re3.empty:
                gb_re3=group.set_index('date')[['num_user']].rename(columns={'num_user':name})
            else:
                gb_re3=gb_re3.join(group.set_index('date')[['num_user']].rename(columns={'num_user':name}), how='outer')
        return gb_re3
    def switch_column_re5():
        # re5 = pd.read_csv('order_backlog_part5.csv') # it's dts.csv
        re5 = pd.read_csv('dts.csv')
        gb = re5.groupby('shipping_date')
        gb_re5 = pd.DataFrame()
        for name, group in gb:
            if gb_re5.empty:
                gb_re5= group.set_index('date')[['num_shipping_date']].rename(columns={'num_shipping_date': name})
            else:
                gb_re5 = gb_re5.join(group.set_index('date')[['num_shipping_date']].rename(columns={'num_shipping_date': name}), how='outer')
        return gb_re5
    # re1=pd.read_csv('order_backlog_part1.csv') # it's total.csv
    re1 = pd.read_csv('total.csv')
    re1=re1.set_index('date')
    re3=switch_column_re3()
    re3.insert(1, 'SL',0)
    re3=re3.fillna(0)
    # re4= pd.read_csv('order_backlog_part4.csv') # it's active.csv
    re4 = pd.read_csv('active.csv')
    re4=re4.set_index('date')
    re5= switch_column_re5()
    # print(re5)
    re5.insert(0, 'DTS1', 0)
    re5.insert(2, 'DTS3', 0)
    re5=re5.fillna(0)
    # re5.insert(3, 'DTS>3', 0)
    # print(re5)
    # print(re1.head())
    ## '2019-08-04'
    # print(re3.head())
    # print(re4.head())
    # print(re5.head())
    # print(re3)
    #----------------------------------------
    g_day=re1.index.tolist()
    row_ex=3
    col_ex=1
    for day in g_day:
        print(day)
        #column_ex =1
        sheet1.cell(row=row_ex, column=col_ex).value=day
        #column_ex=2
        sheet1.cell(row=row_ex, column=col_ex+1).value=re1.loc[day][0]
        #column_ex=3
        sheet1.cell(row=row_ex, column=col_ex+2).value = 0
        try:
            #column_ex=4 - buyer cancel
            sheet1.cell(row=row_ex, column=col_ex+3).value = re3.loc[day][0]
        except: sheet1.cell(row=row_ex, column=col_ex+3).value = 0
        try:
            #column_ex=5 - seller cancel
            sheet1.cell(row=row_ex, column=col_ex+4).value = re3.loc[day][1]
        except: sheet1.cell(row=row_ex, column=col_ex+4).value = 0
        try:
            #column_ex=6 - Auto
            if re3.loc[day][2] == np.NaN:
                sheet1.cell(row=row_ex, column=col_ex + 5).value = 0
            else:
                sheet1.cell(row=row_ex, column=col_ex+5).value = re3.loc[day][2]
        except: sheet1.cell(row=row_ex, column=col_ex+5).value = 0
        #coulmn_ex=7 - Unknown
        try:
            sheet1.cell(row=row_ex, column=col_ex+6).value=re3.loc[day][3]
        except: sheet1.cell(row=row_ex, column=col_ex+6).value=0
        try:
            #column_ex=8 - Active
            sheet1.cell(row=row_ex, column=col_ex+7).value = re4.loc[day][0]
        except: sheet1.cell(row=row_ex, column=col_ex+7).value = 0
        try:
            #column_ex=9 - DTS1
            sheet1.cell(row=row_ex, column=col_ex+8).value = re5.loc[day][0]
        except: sheet1.cell(row=row_ex, column=col_ex+8).value = 0
        try:
            #column_ex=10 - DTS2
            sheet1.cell(row=row_ex, column=col_ex+9).value = re5.loc[day][1]
        except: sheet1.cell(row=row_ex, column=col_ex+9).value =0
        try:
            #column_ex=11 -DTS3
            sheet1.cell(row=row_ex, column=col_ex+10).value = re5.loc[day][2]
        except: sheet1.cell(row=row_ex, column=col_ex+10).value = 0
        try:
            #column_ex=12 -DTS>3
            sheet1.cell(row=row_ex, column=col_ex+11).value = re5.loc[day][3]
        except: sheet1.cell(row=row_ex, column=col_ex+11).value = 0
        row_ex+=1
    wb.save('report.xlsx')
write_to_excel()