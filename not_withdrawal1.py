import pandas as pd
import openpyxl
def write_workbook(workbook, month):
    # list_workbook=['report_1.xlsx','report_2.xlsx','report_3.xlsx']
    wb = openpyxl.load_workbook(workbook)
    sheet1 = wb['Sheet1']
    def set_header():
        header=['USERID', 'SHOPID', 'SHOP_NAME', 'MAIN_CATEGORY', 'ADO_MAY', 'ADO_JUN', 'ADO_JUL',
                'GMV_MAY', 'GMV_JUN', 'GMV_JUL']
        col_name=[chr(char) for char in range(ord('A'), len(header)+ord('A')+1) ]
        row=1
        for col in range(1, len(header)+1):
            sheet1.column_dimensions[col_name[col-1]].width=15
            sheet1.cell(row, col).value=header[col-1]
        wb.save(workbook)
    def input_values(month3):
        df_may=pd.read_csv(month3[0]).fillna('')
        df_jun=pd.read_csv(month3[1]).iloc[:, 4:].fillna('')
        df_jul=pd.read_csv(month3[2]).iloc[:, 4:].fillna('')
        row=2
        row_pd=0
        # print(df_jul.head())
        while True:
            try:
                #userid
                sheet1.cell(row, 1).value = df_may.iloc[row_pd, 0]
                #shop_id
                sheet1.cell(row, 2).value = df_may.iloc[row_pd, 1]
                #shop_name
                sheet1.cell(row, 3).value = df_may.iloc[row_pd, 2]
                #main_category
                sheet1.cell(row, 4).value = df_may.iloc[row_pd, 3]
                #ADO MAY
                sheet1.cell(row, 5).value = df_may.iloc[row_pd, 4]
                #ADO JUNE
                sheet1.cell(row, 6).value = df_jun.iloc[row_pd, 0]
                #ADO JULY
                sheet1.cell(row, 7).value = df_jul.iloc[row_pd, 0]
                #GMV MAY
                sheet1.cell(row, 8).value = df_may.iloc[row_pd, 5]
                #GMV JUNE
                sheet1.cell(row, 9).value = df_jun.iloc[row_pd, 1]
                #GMV JULY
                sheet1.cell(row, 10).value = df_jul.iloc[row_pd, 1]
            except : break
            else:
                row+=1
                row_pd+=1
        wb.save(workbook)
    set_header()
    input_values(month)
# list_workbook=['report_1.xlsx','report_2.xlsx','report_3.xlsx']
# wb_month=[['May_001.csv', 'June_001.csv', 'July_001.csv'],['May_002.csv', 'June_002.csv', 'July_002.csv'],['May_003.csv', 'June_003.csv', 'July_003.csv']]
# for i, item in enumerate(list_workbook):
#     write_workbook(item, wb_month[i])
write_workbook('report_1.xlsx', ['May_001.csv', 'June_001.csv', 'July_001.csv'])
# write_workbook('report_2.xlsx', ['May_002.csv', 'June_002.csv', 'July_002.csv'])
# write_workbook('report_3.xlsx', ['May_003.csv', 'June_003.csv', 'July_003.csv'])