import pandas as pd
import os
import openpyxl
def write_2excel():
    selected_date=pd.date_range(start='11/11/2018', end='09/09/2019')
    seller=('OS', 'MS')
    # wb=openpyxl.Workbook()
    wb=openpyxl.load_workbook('nfr_lsr.xlsx')
    for t in seller:
        print(t)
        sheet=wb[t]
        # wb.create_sheet(t)
        # sheet=wb.get_sheet_by_name(t)
        # sheet=wb[t]
        col = 2
        for d in selected_date:
            if d.day == d.month or (d.day == 29 and d.day == 6):
                d=d.strftime('%Y-%m-%d')
                print(d)
                detail=f"{t}_detail_{d}.csv"
                nfr=f"{t}_NFR_Lessthan3_{d}.csv"
                lsr=f"{t}_LSR_Lessthan3_{d}.csv"

                shop = pd.read_excel('shopid.xlsx', sheet_name=t)
                shop=shop.set_index('shopid')

                df_detail=pd.DataFrame(pd.read_csv(detail))
                df_detail=df_detail.set_index('shopid')
                df_detail=shop.join(df_detail, how='left')

                df_nfr=pd.DataFrame(pd.read_csv(nfr))
                df_nfr=df_nfr[df_nfr[df_nfr.columns[1]] < 0.03]
                df_nfr=df_nfr.set_index('shopid')
                df_nfr=shop.join(df_nfr, how='left')

                df_lsr=pd.DataFrame(pd.read_csv(lsr))
                df_lsr=df_lsr.set_index('shopid')
                df_lsr=shop.join(df_lsr, how='left')

                def put_shopid():
                    s = pd.read_excel('shopid.xlsx', sheet_name=t)
                    # print(s.iloc[2:, 0])
                    column=1
                    for row in range(3,len(s)+3):
                        sheet.cell(row=row, column=column).value=s.iloc[row-3, 0]

                def put_data2date():
                    def header():
                        sheet.cell(row=1, column=col).value=d
                        sheet.cell(row=2, column=col).value='ado'
                        sheet.cell(row=2, column=col+1).value='nfr'
                        sheet.cell(row=2, column=col+2).value='lsr'
                    for row in range(3,len(shop)+3):
                        #ado
                        sheet.cell(row=row, column=col).value=df_detail.iloc[row-3, 0]
                        #nfr
                        sheet.cell(row=row, column=col+1).value=df_nfr.iloc[row-3, 0]
                        #lsr
                        sheet.cell(row=row, column=col+2).value=df_lsr.iloc[row - 3, 0]
                    header()
                put_shopid()
                put_data2date()
                col += 3
            else: continue
    wb.save('nfr_lsr.xlsx')
write_2excel()







