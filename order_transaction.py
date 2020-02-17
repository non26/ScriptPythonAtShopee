import pandas as pd
import openpyxl
def write_workbook():
    wb=openpyxl.load_workbook('withdrawal_order_transaction_report.xlsx')
    sheet2=wb['Order']
    dict_bank_col={}
    def set_header():
        # trans_order=pd.read_csv('transaction_order.csv')
        list_bank=sorted(['SCB','GHBANK','TMB','BBL','TISCO','ISAM','ICBC','LHB','KTB',
                   'BACC', 'KBANK','CIMB', 'THANACHART','BAY','UOB', 'STANCHAR','KIATNAKIN','GSB',
                   'KTC', 'BAAC'], reverse=False)
        for char in range(ord('A'), ord('A')+22):
            sheet2.column_dimensions[chr(char)].width=10
        sheet2.cell(1,1).value='time'
        sheet2.cell(1,2).value='bank'
        index=0
        for col in range(3,len(list_bank)+3):
            sheet2.cell(1,col).value =list_bank[index]
            nonlocal  dict_bank_col
            dict_bank_col[list_bank[index]]=col
            index+=1
        wb.save('withdrawal_order_transaction_report.xlsx')

    def input_value():
        df_bank = sorted(['SCB', 'GHBANK', 'TMB', 'BBL', 'TISCO', 'ISAM', 'ICBC', 'LHB', 'KTB',
                            'BACC', 'KBANK', 'CIMB', 'THANACHART', 'BAY', 'UOB', 'STANCHAR', 'KIATNAKIN', 'GSB',
                            'KTC', 'BAAC', 'GHB', 'ิBBL', 'LH', 'LHBANK'], reverse=False)
        trans_order = pd.read_csv('transaction_order.csv')
        unique_time = pd.Series(trans_order['time'].unique())
        start = 0
        i = 0
        row_for_trans = 2
        row_for_amount = 3
        row = 0
        while True:
            row_pd = 0
            while True:
                try:
                    if trans_order.iloc[row, 0] == unique_time[i]:
                        row += 1
                        continue
                    else:
                        sheet2.cell(row_for_trans, 1).value = unique_time[i]
                        sheet2.cell(row_for_trans, 2).value = 'total transaction'
                        sheet2.cell(row_for_amount, 2).value = 'GMV'
                        i += 1
                        row_pd = row
                        break
                except Exception:
                    break
            # print(row_pd)
            same_date = trans_order.iloc[start:row_pd, :]
            if same_date.empty: break
            start = row_pd
            # print(same_date)
            index_for_same_date = 0
            l = 0
            # adapt sheet width
            for item in range(ord('C'), ord('C')+len(df_bank)-4 ):
                sheet2.column_dimensions[chr(item)].width = 25
            # put total transaction, gmv that corresponding to its bank
            while True:
                try:
                    x=list(same_date.iloc[index_for_same_date, 1])
                    while True:
                        try:
                            y=x.index(' ')
                            del x[y]
                        except Exception : break
                    for item in x:
                        if ord('a')<=ord(item)<=ord('z') or ord('A')<=ord(item)<=ord('Z'):pass
                        else: x.remove('item')
                    bank = ''.join(x).upper()
                    # print(bank)
                    if bank in df_bank: #include unclean characters
                        if bank=='GHB':
                            # total transaction
                            sheet2.cell(row_for_trans,dict_bank_col['GHBANK']).value = float(same_date.iloc[index_for_same_date, 2])
                            # total GMV
                            sheet2.cell(row_for_amount,dict_bank_col['GHBANK']).value = float(same_date.iloc[index_for_same_date, 3])
                            l += 1
                            index_for_same_date += 1
                        # elif bank==  'ิBBL':
                        #     # total transaction
                        #     sheet2.cell(row_for_trans, dict_bank_col['BBL']).value = float(same_date.iloc[index_for_same_date, 2])
                        #     # total GMV
                        #     sheet2.cell(row_for_amount, dict_bank_col['BBL']).value = float(same_date.iloc[index_for_same_date, 3])
                        #     l += 1
                        #     index_for_same_date += 1
                        elif bank in ('LH','LHBANK') :
                            # total transaction
                            sheet2.cell(row_for_trans, dict_bank_col['LHB']).value = float(same_date.iloc[index_for_same_date, 2])
                            # total GMV
                            sheet2.cell(row_for_amount, dict_bank_col['LHB']).value = float(same_date.iloc[index_for_same_date, 3])
                            l += 1
                            index_for_same_date += 1
                        else:
                            # total transaction
                            sheet2.cell(row_for_trans, dict_bank_col[bank]).value = float(same_date.iloc[index_for_same_date, 2])
                            # total gmv
                            sheet2.cell(row_for_amount,dict_bank_col[bank]).value = float(same_date.iloc[index_for_same_date, 3])
                            l += 1
                            index_for_same_date += 1
                    else:
                        index_for_same_date = l
                        l += 1
                        continue
                except Exception:break
            row_for_amount += 2
            row_for_trans += 2
            wb.save('withdrawal_order_transaction_report.xlsx')
    set_header()
    input_value()
write_workbook()






