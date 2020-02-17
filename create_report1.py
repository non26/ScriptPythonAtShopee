import pandas as pd
import openpyxl
from datetime import datetime
from datetime import date
def write_to_excel():
    wb=openpyxl.load_workbook('report.xlsx')
    sheet1=wb['Sheet1']
    def header():
        #row1
        sheet1.cell(row=1, column=4).value = 'Reason'
        sheet1.cell(row=1, column=5).value = 'Pay_result'
        column_ex=6
        for i in pd.date_range(start='2019-09-13', end=date.today().strftime('%Y-%m-%d')):
            sheet1.cell(row=1, column=column_ex).value= i.date()
            column_ex += 1
        wb.save('report.xlsx')
    def airpay():
        def header():
            sheet1.cell(row=2, column=1).value = 'AirPay Wallet V.2'
            sheet1.cell(row=2, column=2).value = '#payment ID'
            sheet1.cell(row=2, column=3).value = 'success'
            sheet1.cell(row=2, column=4).value = 'success'
            sheet1.cell(row=2, column=5).value = 'success'
            reason = [('Network issues or timeout', 101), ('account locked',221),
                      ('wrong wallet PIN',226), ('no stuch txn',501),(r"AP cannot return whether success/fail",521),
                      ('insufficient wallet balance',522), ('exceed payment limit',527), ('bank maintenance',530),
                      ('insufficient BA balance',531), ('abnormal BA status',532), ('payment cancel order expired/cancelled','no pay_result'),
                      ('others', '')]
            row=3
            for i, j in enumerate(reason):
                col = 3
                sheet1.cell(row=row, column=col).value='fail'
                col+=1
                sheet1.cell(row=row, column=col).value=j[0]
                col+=1
                sheet1.cell(row=row, column=col).value = j[1]
                row+=1
        def header_for_other():
            #start row = 14, col=5
            dict_row = {101: 3, 221: 4, 226: 5, 501: 6, 521: 7, 522: 8, 527: 9, 530: 10, 531: 11, 532: 12, 'NON': 13}
            gb = pd.DataFrame(pd.read_csv('airpay_fail.csv'))
            gb = gb.fillna('NON')
            diff_other=list(set(gb['result'].unique()).difference(set(dict_row.keys())))
            diff_other.sort(key=lambda x:str(x))
            row=14
            col=5
            for n in range(1, len(diff_other) + 1):
                dict_row[diff_other[n - 1]]=row
                sheet1.cell(row=row, column=col).value=diff_other[n - 1]
                row+=1
            return dict_row
        def success():
            # sucess row=2, col=6
            row, col = 2, 6
            row_pd, col_pd = 0, 1
            df = pd.read_csv('airpay_success.csv').iloc[:, 1:]
            df = df.set_index('DATE')
            d = pd.DataFrame({'DATE': pd.date_range(start='2019-09-13', end=date.today().strftime('%Y-%m-%d')), 'num': 0.0})
            d = d.set_index('DATE')
            df1=d.join(df)
            while True:
                try:
                    sheet1.cell(row=row, column=col).value=df1.iloc[row_pd, col_pd]
                    row_pd+=1
                    col+=1
                except: break
        def fail(matching):
            #fail row=3, col=6
            # row, col = 3,6
            gb=pd.read_csv('airpay_fail.csv').groupby(by='day')
            d = pd.Series(pd.date_range(start='2019-09-13', end=date.today().strftime('%Y-%m-%d')))
            d = d.dt.strftime('%Y-%m-%d')
            dict_row = matching
            col=6
            r=0
            for name, group in gb:
                row_pd = 0
                group=group.fillna('NON')
                # print(group)
                while True:
                    if name == d[r]:
                        try:
                            sheet1.cell(row=dict_row[group.iloc[row_pd, 2]], column=col).value=group.iloc[row_pd,3]
                            row_pd+=1
                        except:
                            col+=1
                            r+=1
                            break
                    else:
                        r+=1
                        col+=1
                        continue
        print('header')
        header()
        print('header for other')
        matching_row=header_for_other()
        print('writing success Airpay')
        success()
        print('writing fail Airpay')
        fail(matching_row)
        wb.save('report.xlsx')
    def giro():
        dict_row = {101: 3, 221: 4, 226: 5, 501: 6, 521: 7, 522: 8, 527: 9, 530: 10, 531: 11, 532: 12,
                'NON':13}
        gb = pd.DataFrame(pd.read_csv('giro_fail.csv'))
        gb = gb.fillna('NON')
        diff_other=list(set(gb['result'].unique()).difference(set(dict_row.keys())))
        added_row = len(diff_other)
        def header():
            x=16+added_row
            sheet1.cell(row=x, column=1).value = 'Giro'
            sheet1.cell(row=x, column=2).value = '#payment ID'
            sheet1.cell(row=x, column=3).value = 'success'
            sheet1.cell(row=x, column=4).value = 'success'
            sheet1.cell(row=x, column=5).value = 'success'
            #-----
            reason = [('Network issues or timeout', 101), ('account locked', 221),
                      ('wrong wallet PIN', 226), ('no stuch txn', 501), (r"AP cannot return whether success/fail", 521),
                      ('insufficient wallet balance', 522), ('exceed payment limit', 527), ('bank maintenance', 530),
                      ('insufficient BA balance', 531), ('abnormal BA status', 532),
                      ('payment cancel order expired/cancelled', 'no pay_result'),
                      ('others', '')]
            row = 17+added_row
            for i, j in enumerate(reason):
                col = 3
                sheet1.cell(row=row, column=col).value = 'fail'
                col += 1
                sheet1.cell(row=row, column=col).value = j[0]
                col += 1
                sheet1.cell(row=row, column=col).value = j[1]
                row += 1
        def header_for_other():
            #start row = 14, col=5
            dict_row = {101: 3, 221: 4, 226: 5, 501: 6, 521: 7, 522: 8, 527: 9, 530: 10, 531: 11, 532: 12, 'NON': 13}
            for key in dict_row:
                dict_row[key]=dict_row[key]+14+added_row
            gb = pd.DataFrame(pd.read_csv('giro_fail.csv'))
            gb = gb.fillna('NON')
            diff_other=list(set(gb['result'].unique()).difference(set(dict_row.keys())))
            diff_other.sort(key=lambda x:str(x))
            row= 14+14+added_row
            col=5
            for n in range(1, len(diff_other) + 1):
                dict_row[diff_other[n - 1]]=row
                sheet1.cell(row=row, column=col).value=diff_other[n - 1]
                row+=1
            return dict_row
        def success():
            # sucess row=16, col=6
            row, col = 16+added_row, 6
            row_pd, col_pd = 0, 1
            df = pd.read_csv('giro_success.csv').iloc[:, 1:]
            df = df.set_index('date')
            d = pd.DataFrame({'DATE': pd.date_range(start='2019-09-13', end=date.today().strftime('%Y-%m-%d')), 'num': 0.0})
            d = d.set_index('DATE')
            df1 = d.join(df)
            while True:
                try:
                    sheet1.cell(row=row, column=col).value = df1.iloc[row_pd, col_pd]
                    row_pd += 1
                    col += 1
                except:
                    break
        def fail(matching):
            # fail row=17+added_row(), col=6
            # row, col = 17+added_row, 6
            dict_row=matching
            # for key in dict_row:
            #     dict_row[key]=dict_row[key]+row
            gb = pd.read_csv('giro_fail.csv').groupby(by='date')
            d = pd.Series(pd.date_range(start='2019-09-13', end=date.today().strftime('%Y-%m-%d')))
            d = d.dt.strftime('%Y-%m-%d')
            col=6
            r=0
            for name, group in gb:
                row_pd = 0
                group=group.fillna('NON')
                while True:
                    if name == d[r]:
                        try:
                            sheet1.cell(row=dict_row[group.iloc[row_pd, 2]], column=col).value=group.iloc[row_pd,3]
                            row_pd+=1
                        except:
                            col+=1
                            r+=1
                            break
                    else:
                        r+=1
                        col+=1
                        continue
        print('header')
        header()
        print('header for other')
        row_matching=header_for_other()
        print('writing success Giro')
        success()
        print('writing fail Giro')
        fail(row_matching)
        wb.save('report.xlsx')
    print('Header Excel')
    header()
    print('Airpay .... ')
    airpay()
    print('Giro .... ')
    giro()
write_to_excel()
